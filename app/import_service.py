import base64
from datetime import date
from io import BytesIO

from .auth_service import require_auth_user
from .collections import normalize_number_fields
from .config import DB_LOCK
from .db import db_execute, get_connection, insert_and_get_id, query_all, query_one
from .errors import ApiError

SHEET_ALIASES = {
    "courses": "courses",
    "course": "courses",
    "disciplines": "courses",
    "discipline": "courses",
    "teachers": "teachers",
    "teacher": "teachers",
    "rooms": "rooms",
    "room": "rooms",
    "groups": "groups",
    "group": "groups",
    "sections": "sections",
    "section": "sections",
}

COURSE_HEADERS = {
    "code": "code",
    "course_code": "code",
    "код": "code",
    "course_code": "code",
    "name": "name",
    "course_name": "name",
    "название": "name",
    "атауы": "name",
    "study_year": "year",
    "year": "year",
    "курс": "year",
    "год": "year",
    "semester": "semester",
    "семестр": "semester",
    "department": "department",
    "faculty": "department",
    "faculty_institute": "department",
    "faculty_or_institute": "department",
    "факультет": "department",
    "факультет_институт": "department",
    "институт": "department",
    "instructor": "instructor_name",
    "teacher": "instructor_name",
    "teacher_name": "instructor_name",
    "instructor_name": "instructor_name",
    "преподаватель": "instructor_name",
    "оқытушы": "instructor_name",
    "programme": "programme",
    "programme_name": "programme",
    "program_name": "programme",
    "program": "programme",
    "образовательная_программа": "programme",
    "бағдарлама": "programme",
    "description": "description",
    "описание": "description",
    "сипаттама": "description",
}

TEACHER_HEADERS = {
    "name": "name",
    "full_name": "name",
    "фио": "name",
    "аты-жөні": "name",
    "email": "email",
    "phone": "phone",
    "телефон": "phone",
    "specialization": "department",
    "faculty": "department",
    "department": "department",
    "faculty_institute": "department",
    "faculty_or_institute": "department",
    "факультет": "department",
    "факультет_институт": "department",
    "институт": "department",
    "специализация": "department",
    "мамандығы": "department",
    "max_hours_per_week": "weekly_hours_limit",
    "max_hours": "weekly_hours_limit",
    "максимум_часов_в_неделю": "weekly_hours_limit",
    "апталық_сағат_лимиті": "weekly_hours_limit",
}

ROOM_HEADERS = {
    "number": "number",
    "room_number": "number",
    "номер": "number",
    "нөмір": "number",
    "capacity": "capacity",
    "вместимость": "capacity",
    "сыйымдылығы": "capacity",
    "building": "building",
    "здание": "building",
    "ғимарат": "building",
    "type": "type",
    "тип": "type",
    "түрі": "type",
    "department": "department",
    "faculty": "department",
    "faculty_institute": "department",
    "faculty_or_institute": "department",
    "факультет": "department",
    "факультет_институт": "department",
    "институт": "department",
    "available": "available",
    "is_available": "available",
    "доступно": "available",
    "қолжетімді": "available",
    "equipment": "equipment",
    "оборудование": "equipment",
    "жабдықтар": "equipment",
}

GROUP_HEADERS = {
    "name": "name",
    "group_name": "name",
    "group_number": "name",
    "номер_группы": "name",
    "топ_нөмірі": "name",
    "student_count": "student_count",
    "students_count": "student_count",
    "количество_студентов": "student_count",
    "студент_саны": "student_count",
    "has_subgroups": "has_subgroups",
    "subgroups": "has_subgroups",
    "подгруппы": "has_subgroups",
}

SECTION_HEADERS = {
    "course_code": "course_code",
    "code": "course_code",
    "код_курса": "course_code",
    "group_name": "group_name",
    "group_number": "group_name",
    "номер_группы": "group_name",
    "топ_нөмірі": "group_name",
    "classes_count": "classes_count",
    "class_count": "classes_count",
    "количество_занятий": "classes_count",
    "сабақ_саны": "classes_count",
    "lesson_type": "lesson_type",
    "type": "lesson_type",
    "тип_занятия": "lesson_type",
    "сабақ_түрі": "lesson_type",
}

REQUIRED_FIELDS = {
    "courses": ["code", "name", "year", "semester", "programme", "department"],
    "teachers": ["name", "email"],
    "rooms": ["number", "capacity", "department"],
    "groups": ["name", "student_count"],
    "sections": ["course_code", "group_name", "classes_count"],
}

ROOM_TYPE_ALIASES = {
    "lecture": "lecture",
    "lecturehall": "lecture",
    "lecture hall": "lecture",
    "лекция": "lecture",
    "лекционный": "lecture",
    "лекционная аудитория": "lecture",
    "practical": "practical",
    "practicalroom": "practical",
    "practical room": "practical",
    "practice": "practical",
    "практика": "practical",
    "практический": "practical",
    "практикалық": "practical",
    "lab": "lab",
    "laboratory": "lab",
    "лаборатория": "lab",
    "зертхана": "lab",
    "seminar": "seminar",
    "семинар": "seminar",
}

TEMPLATE_HEADERS = {
    "Disciplines": [
        "code",
        "name",
        "year",
        "semester",
        "programme",
        "department",
        "instructor_name",
        "description",
    ],
    "Teachers": ["name", "email", "phone", "department"],
    "Rooms": ["number", "capacity", "building", "type", "department", "available", "equipment"],
    "Groups": ["name", "student_count", "has_subgroups"],
    "Sections": ["course_code", "group_name", "classes_count", "lesson_type"],
}

TEMPLATE_ROWS = {
    "Disciplines": [
        [
            "CS101",
            "Programming 1",
            date.today().year,
            1,
            "Программная инженерия (6B06101)",
            "Факультет компьютерных систем и профессионального образования (КСиПО-БжЦТ)",
            "Aruzhan Saparova",
            "Introduction to programming",
        ],
    ],
    "Teachers": [
        [
            "Aruzhan Saparova",
            "aruzhan@kazatu.edu.kz",
            "+7 777 000 00 00",
            "Факультет компьютерных систем и профессионального образования (КСиПО-БжЦТ)",
        ],
    ],
    "Rooms": [
        [
            "101",
            30,
            "Main Building",
            "lecture",
            "Факультет компьютерных систем и профессионального образования (КСиПО-БжЦТ)",
            "yes",
            "Projector, whiteboard",
        ],
    ],
    "Groups": [
        ["SE-23-01", 24, "yes"],
    ],
    "Sections": [
        ["CS101", "SE-23-01", 2, "lecture"],
    ],
}

AVAILABLE_ALIASES = {
    "1": 1,
    "true": 1,
    "yes": 1,
    "available": 1,
    "да": 1,
    "иә": 1,
    "нет": 0,
    "no": 0,
    "false": 0,
    "0": 0,
    "not_available": 0,
    "not available": 0,
    "жоқ": 0,
}


def _load_workbook(file_bytes):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ApiError(
            500,
            "internal_server_error",
            "Excel import dependency is not installed on the server.",
        ) from exc

    try:
        return load_workbook(filename=BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ApiError(
            400,
            "bad_request",
            "Не удалось прочитать Excel файл. Используйте формат .xlsx.",
        ) from exc


def _decode_excel_payload(payload):
    file_name = (payload.get("fileName") or "").strip()
    file_content = payload.get("fileContent")

    if not file_name or not file_content:
        raise ApiError(
            400,
            "fill_required_fields",
            "Заполните поля: fileName, fileContent",
            {"fields": ["fileName", "fileContent"]},
        )

    if not file_name.lower().endswith(".xlsx"):
        raise ApiError(
            400,
            "bad_request",
            "Поддерживаются только Excel файлы формата .xlsx.",
        )

    if "," in file_content:
        file_content = file_content.split(",", 1)[1]

    try:
        return base64.b64decode(file_content)
    except Exception as exc:
        raise ApiError(400, "bad_request", "Некорректное содержимое файла.") from exc


def _normalize_header(value):
    if value is None:
        return ""
    return str(value).strip().lower().replace("\n", " ").replace("-", "_")


def _normalize_sheet_name(sheet_name):
    normalized = _normalize_header(sheet_name).replace(" ", "")
    return SHEET_ALIASES.get(normalized)


def _normalize_room_type(value):
    if value in (None, ""):
        return ""
    normalized = _normalize_header(value).replace("_", " ")
    compact = normalized.replace(" ", "")
    return (
        ROOM_TYPE_ALIASES.get(compact)
        or ROOM_TYPE_ALIASES.get(normalized)
        or str(value).strip().lower()
    )


def _normalize_availability(value):
    if value in (None, ""):
        return 1
    if isinstance(value, bool):
        return 1 if value else 0
    normalized = _normalize_header(value).replace("_", " ")
    compact = normalized.replace(" ", "_")
    return AVAILABLE_ALIASES.get(compact, AVAILABLE_ALIASES.get(normalized, 1 if value else 0))


def _read_sheet_rows(sheet, header_aliases):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    raw_headers = rows[0]
    canonical_headers = []
    for raw_header in raw_headers:
        normalized = _normalize_header(raw_header)
        canonical_headers.append(header_aliases.get(normalized, normalized))

    parsed_rows = []
    for row_index, row in enumerate(rows[1:], start=2):
        if not any(cell not in (None, "") for cell in row):
            continue

        parsed = {}
        for column_index, value in enumerate(row):
            if column_index >= len(canonical_headers):
                continue
            header = canonical_headers[column_index]
            if not header:
                continue
            parsed[header] = value.strip() if isinstance(value, str) else value
        parsed_rows.append((row_index, parsed))

    return parsed_rows


def _validate_required_fields(entity_name, row_index, payload):
    missing = [
        field for field in REQUIRED_FIELDS[entity_name] if payload.get(field) in (None, "")
    ]
    if missing:
        raise ApiError(
            400,
            "bad_request",
            f"Лист {entity_name}: строка {row_index}. Отсутствуют поля: {', '.join(missing)}.",
        )


def _upsert_course(connection, payload):
    normalized = normalize_number_fields(payload, ["year", "study_year", "semester"])
    instructor_name = (normalized.get("instructor_name") or "").strip()
    instructor_id = None
    if instructor_name:
        teacher = query_one(
            connection,
            """
            SELECT id, name
            FROM teachers
            WHERE lower(name) = lower(?)
            """,
            (instructor_name,),
        )
        if teacher:
            instructor_id = teacher["id"]
            instructor_name = teacher["name"]

    existing = query_one(
        connection,
        "SELECT id FROM courses WHERE lower(code) = lower(?)",
        (normalized["code"],),
    )
    if existing:
        db_execute(
            connection,
            """
            UPDATE courses
            SET
                name = ?,
                code = ?,
                description = ?,
                year = ?,
                semester = ?,
                department = ?,
                instructor_id = ?,
                instructor_name = ?,
                programme = ?
            WHERE id = ?
            """,
            (
                normalized["name"],
                normalized["code"],
                normalized.get("description", "") or "",
                normalized.get("year", normalized.get("study_year")),
                normalized.get("semester"),
                normalized.get("department", "") or "",
                instructor_id,
                instructor_name,
                normalized.get("programme", normalized.get("programme_name", "")) or "",
                existing["id"],
            ),
        )
        return "updated"

    insert_and_get_id(
        connection,
        """
        INSERT INTO courses (
            name, code, credits, hours, description,
            year, semester, department, instructor_id, instructor_name, programme
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            normalized["name"],
            normalized["code"],
            None,
            None,
            normalized.get("description", "") or "",
            normalized.get("year", normalized.get("study_year")),
            normalized.get("semester"),
            normalized.get("department", "") or "",
            instructor_id,
            instructor_name,
            normalized.get("programme", normalized.get("programme_name", "")) or "",
        ),
    )
    return "inserted"


def _upsert_teacher(connection, payload):
    normalized = normalize_number_fields(payload, ["weekly_hours_limit", "max_hours_per_week"])
    existing = query_one(
        connection,
        "SELECT id FROM teachers WHERE lower(email) = lower(?)",
        (normalized["email"],),
    )
    if existing:
        db_execute(
            connection,
            """
            UPDATE teachers
            SET name = ?, email = ?, phone = ?, department = ?, weekly_hours_limit = ?
            WHERE id = ?
            """,
            (
                normalized["name"],
                normalized["email"],
                normalized.get("phone", "") or "",
                normalized.get("department", normalized.get("specialization", "")) or "",
                normalized.get("weekly_hours_limit", normalized.get("max_hours_per_week")),
                existing["id"],
            ),
        )
        return "updated"

    insert_and_get_id(
        connection,
        """
        INSERT INTO teachers (name, email, phone, department, weekly_hours_limit)
        VALUES (?, ?, ?, ?, ?)
        """,
        (
            normalized["name"],
            normalized["email"],
            normalized.get("phone", "") or "",
            normalized.get("department", normalized.get("specialization", "")) or "",
            normalized.get("weekly_hours_limit", normalized.get("max_hours_per_week")),
        ),
    )
    return "inserted"


def _upsert_room(connection, payload):
    normalized = normalize_number_fields(payload, ["capacity"])
    normalized["type"] = _normalize_room_type(normalized.get("type"))
    normalized["available"] = _normalize_availability(normalized.get("available"))
    existing = query_one(
        connection,
        "SELECT id FROM rooms WHERE number = ?",
        (str(normalized["number"]),),
    )
    if existing:
        db_execute(
            connection,
            """
            UPDATE rooms
            SET number = ?, capacity = ?, building = ?, type = ?, equipment = ?, department = ?, available = ?
            WHERE id = ?
            """,
            (
                str(normalized["number"]),
                normalized["capacity"],
                normalized.get("building", "") or "",
                normalized.get("type", "") or "",
                normalized.get("equipment", "") or "",
                normalized.get("department", "") or "",
                normalized.get("available", 1),
                existing["id"],
            ),
        )
        return "updated"

    insert_and_get_id(
        connection,
        """
        INSERT INTO rooms (number, capacity, building, type, equipment, department, available)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            str(normalized["number"]),
            normalized["capacity"],
            normalized.get("building", "") or "",
            normalized.get("type", "") or "",
            normalized.get("equipment", "") or "",
            normalized.get("department", "") or "",
            normalized.get("available", 1),
        ),
    )
    return "inserted"


def _normalize_bool(value):
    if value in (None, ""):
        return 0
    if isinstance(value, bool):
        return 1 if value else 0
    normalized = _normalize_header(value).replace("_", " ")
    if normalized in {"1", "true", "yes", "да", "иә", "a/b", "a / b"}:
        return 1
    return 0


def _upsert_group(connection, payload):
    normalized = normalize_number_fields(payload, ["student_count"])
    has_subgroups = _normalize_bool(payload.get("has_subgroups"))
    existing = query_one(
        connection,
        "SELECT id FROM groups WHERE lower(name) = lower(?)",
        (normalized["name"],),
    )
    if existing:
        db_execute(
            connection,
            """
            UPDATE groups
            SET name = ?, student_count = ?, has_subgroups = ?
            WHERE id = ?
            """,
            (
                normalized["name"],
                normalized["student_count"],
                has_subgroups,
                existing["id"],
            ),
        )
        return "updated"

    insert_and_get_id(
        connection,
        """
        INSERT INTO groups (name, student_count, has_subgroups)
        VALUES (?, ?, ?)
        """,
        (
            normalized["name"],
            normalized["student_count"],
            has_subgroups,
        ),
    )
    return "inserted"


def _upsert_section(connection, payload):
    normalized = normalize_number_fields(payload, ["classes_count"])
    course = query_one(
        connection,
        """
        SELECT id, name, code
        FROM courses
        WHERE lower(code) = lower(?)
        """,
        (normalized["course_code"],),
    )
    if not course:
        raise ApiError(
            400,
            "bad_request",
            f"Для секции не найден курс с кодом '{normalized['course_code']}'.",
        )

    group = query_one(
        connection,
        """
        SELECT id, name
        FROM groups
        WHERE lower(name) = lower(?)
        """,
        (normalized["group_name"],),
    )
    if not group:
        raise ApiError(
            400,
            "bad_request",
            f"Для секции не найдена группа '{normalized['group_name']}'.",
        )

    existing = query_one(
        connection,
        "SELECT id FROM sections WHERE course_id = ? AND group_id = ?",
        (course["id"], group["id"]),
    )
    if existing:
        db_execute(
            connection,
            """
            UPDATE sections
            SET course_id = ?, course_name = ?, group_id = ?, group_name = ?, classes_count = ?, lesson_type = ?
            WHERE id = ?
            """,
            (
                course["id"],
                course["name"],
                group["id"],
                group["name"],
                normalized["classes_count"],
                normalized.get("lesson_type", "lecture") or "lecture",
                existing["id"],
            ),
        )
        return "updated"

    insert_and_get_id(
        connection,
        """
        INSERT INTO sections (course_id, course_name, group_id, group_name, classes_count, lesson_type)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            course["id"],
            course["name"],
            group["id"],
            group["name"],
            normalized["classes_count"],
            normalized.get("lesson_type", "lecture") or "lecture",
        ),
    )
    return "inserted"


def import_excel_data(headers, payload):
    user = require_auth_user(headers)
    if user["role"] != "admin":
        raise ApiError(403, "forbidden", "Недостаточно прав")

    workbook = _load_workbook(_decode_excel_payload(payload))
    sheet_map = {
        "courses": COURSE_HEADERS,
        "teachers": TEACHER_HEADERS,
        "rooms": ROOM_HEADERS,
        "groups": GROUP_HEADERS,
        "sections": SECTION_HEADERS,
    }
    recognized_sheets = []
    parsed_sheets = {}
    summary = {
        "courses": {"inserted": 0, "updated": 0},
        "teachers": {"inserted": 0, "updated": 0},
        "rooms": {"inserted": 0, "updated": 0},
        "groups": {"inserted": 0, "updated": 0},
        "sections": {"inserted": 0, "updated": 0},
    }

    with DB_LOCK:
        with get_connection() as connection:
            for sheet in workbook.worksheets:
                entity_name = _normalize_sheet_name(sheet.title)
                if not entity_name:
                    continue

                recognized_sheets.append(sheet.title)
                parsed_sheets[entity_name] = _read_sheet_rows(sheet, sheet_map[entity_name])

            # Teachers must be imported before courses so instructor_name can resolve to instructor_id.
            for entity_name in ("teachers", "courses", "rooms", "groups", "sections"):
                rows = parsed_sheets.get(entity_name, [])
                for row_index, row_payload in rows:
                    _validate_required_fields(entity_name, row_index, row_payload)
                    if entity_name == "courses":
                        result = _upsert_course(connection, row_payload)
                    elif entity_name == "teachers":
                        result = _upsert_teacher(connection, row_payload)
                    elif entity_name == "rooms":
                        result = _upsert_room(connection, row_payload)
                    elif entity_name == "groups":
                        result = _upsert_group(connection, row_payload)
                    else:
                        result = _upsert_section(connection, row_payload)
                    summary[entity_name][result] += 1

            if not recognized_sheets:
                raise ApiError(
                    400,
                    "bad_request",
                    "В Excel не найдены листы Disciplines, Teachers, Rooms, Groups или Sections.",
                )

            connection.commit()

    total_inserted = sum(item["inserted"] for item in summary.values())
    total_updated = sum(item["updated"] for item in summary.values())
    return {
        "message": "Excel import completed successfully.",
        "recognizedSheets": recognized_sheets,
        "summary": summary,
        "totals": {
            "inserted": total_inserted,
            "updated": total_updated,
        },
    }


def generate_import_template(headers):
    user = require_auth_user(headers)
    if user["role"] != "admin":
        raise ApiError(403, "forbidden", "Недостаточно прав")

    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ApiError(
            500,
            "internal_server_error",
            "Excel import dependency is not installed on the server.",
        ) from exc

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet_name, headers_row in TEMPLATE_HEADERS.items():
        sheet = workbook.create_sheet(title=sheet_name)
        sheet.append(headers_row)
        for row in TEMPLATE_ROWS[sheet_name]:
            sheet.append(row)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def generate_schedule_export(headers):
    user = require_auth_user(headers)
    if user["role"] != "admin":
        raise ApiError(403, "forbidden", "Недостаточно прав")

    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ApiError(
            500,
            "internal_server_error",
            "Excel export dependency is not installed on the server.",
        ) from exc

    with DB_LOCK:
        with get_connection() as connection:
            schedules = query_all(
                connection,
                """
                SELECT
                    course_name,
                    group_name,
                    subgroup,
                    teacher_name,
                    room_number,
                    day,
                    start_hour,
                    semester,
                    year,
                    algorithm
                FROM schedules
                ORDER BY day, start_hour, course_name, group_name, id
                """,
            )

    if not schedules:
        raise ApiError(400, "bad_request", "Расписание ещё не сгенерировано.")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Schedule"
    sheet.append(
        [
            "course_name",
            "group_name",
            "subgroup",
            "teacher_name",
            "room_number",
            "day",
            "start_hour",
            "semester",
            "year",
            "algorithm",
        ]
    )

    for item in schedules:
        sheet.append(
            [
                item.get("course_name", ""),
                item.get("group_name", ""),
                item.get("subgroup", ""),
                item.get("teacher_name", ""),
                item.get("room_number", ""),
                item.get("day", ""),
                item.get("start_hour", ""),
                item.get("semester", ""),
                item.get("year", ""),
                item.get("algorithm", ""),
            ]
        )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
